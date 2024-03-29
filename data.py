import math
import numpy as np
import yfinance as yf
import pandas as pd
from zoneinfo import ZoneInfo
import pickle
from scipy.stats import norm
import datetime
from scipy import interpolate
from scipy.optimize import root_scalar
from openpyxl import load_workbook
import os
from .impvol import implied_vol
from .interest import NelsonSiegel


class OptionData:
    """Class for loading options data from Yahoo Finance.

    Attributes:
        ticker (str): Ticker symbol.
        access_time (Timestamp): Date and time at which the data is loaded from Yahoo Finance.
        underlying_price (float): Current underlying asset.
        forward_prices (DataFrame): Pandas DataFrame containing the synthetic forward prices
            corresponding to different option maturities. The index of the DataFrame is the maturity
            date and the columns are:
                timeToMaturity (float): Time to maturity in years.
                discountFactor (float): Discount factor between access_time and this to maturity.
                forwardPrice (float): Forward price of the underlying asset for this maturity.
        options (DataFrame): Pandas DataFrame containing the options data. The index of the
            DataFrame is a multi-index with levels 'maturity' (date as a Timestamp), type
                ('P' or 'C' for put/call) and strike (float). The columns are:
                price (float): Option mid-price.
                spread (float): Bid-ask spread.
                volume (float): Traded volume.
                logStrike (float): log(strike/forward), i.e. the log-moneyness of a put option.
                OTM (bool): True if out-of-the-money or at-the-money, False if in-the-money.
                impliedVol (float): Implied volatility.
                totalVar (float): Total variance.
                delta (float): Delta.
            Note that not all the columns may be present (e.g. impliedVol and totalVar are added 
            after calling `process_data`).

    Methods:
        load_from_yahoo: Loads options data from Yahoo Finance.
        process_data: Computes forward prices and implied volatilities.
        save_to_disk: Saves the object to disk in pickle or Excel format.
        load_from_disk: Loads the object from disk.

    Notes:
        Time in years is measured as the number of calendar days / 365. This needs to be changed in
        the future.
    """

    ticker: str
    access_time: pd.Timestamp
    underlying_price: float
    forward_prices: pd.DataFrame
    options: pd.DataFrame

    @classmethod
    def load_from_yahoo(cls, ticker, min_maturity=None, max_maturity=None, access_time=None):
        """Loads options data from Yahoo Finance.

        This method populates the object with the options data from Yahoo Finance. Implied 
        volatilities and forward prices are not computed at this stage.

        Args:
            ticker (str): Ticker symbol.
            min_maturity (str | float): Minimum maturity of the options to load. If a string, it
                must be in the format YYYY-MM-DD. If a float, must be time in years.
            max_maturity  (str | float): Maximum maturity of the options to load. If a string, it
                must be in the format YYYY-MM-DD. If a float, must be time in years.
            expiration_time (str): Time of day at which options expire. Must be in the format HH:MM.
            expiration_timezone (str) : Timezone of `expiration_time`.
            access_time (datetime): Time at which the data is accessed. If None, the current time is
                used. This is useful if you download data when the market is closed; in this case
                set `access_time` to the time at which the market closed. If `access_time` does not
                have timezone information, it is assumed to be in the same as `expiration_timezone`.

        Returns:
            YFinanceData object
        """
        # Empty object which we are going to fill with data
        data = cls()
        data.ticker = ticker

        yf_ticker = yf.Ticker(ticker)
        try:
            data.underlying_price = (yf_ticker.info['bid'] + yf_ticker.info['ask']) / 2
        except KeyError:
            data.underlying_price = math.nan

        # Set time at which the data is loaded.
        # The `.astimezone()` adds timezone information to the datetime object if it does not have
        # it (and does nothing if it already has it); the timezone is set to the system timezone by
        # default.
        # We need to be careful with timezone information when subtracting dates and times
        # (e.g. for computing time to maturity) for short options. In principle, for options with
        # not so short maturities this could be ignored.
        timezone = yf_ticker.info['timeZoneFullName'] 
        if access_time is None:
            data.access_time = datetime.datetime.now(tz=ZoneInfo(timezone))
        else:
            data.access_time = access_time.astimezone(ZoneInfo(timezone))

        # Loading option data from Yahoo 
        # First we load each maturity to a separate DataFrame as returned by yfinance, then join
        options_dataframes = []

        # The following will be a dictionary mapping maturities as Pandas Timestamp to time to
        # maturities in years
        time_to_maturities = {}
        for maturity_str in yf_ticker.options:      # yfinance stores maturities as strings
            maturity = datetime.datetime.strptime(maturity_str, "%Y-%m-%d").date()

            # Skip options with maturities out of the range specified by the user (if any)
            if min_maturity is not None:
                if isinstance(min_maturity, str):
                    if datetime.datetime.strptime(min_maturity, "%Y-%m-%d").date() > maturity:
                        continue
                else:
                    if min_maturity > (maturity - data.access_time.date()).days / 365:
                        continue
            if max_maturity is not None:
                if isinstance(max_maturity, str):
                    if datetime.datetime.strptime(max_maturity, "%Y-%m-%d").date() < maturity:
                        continue
                else:
                    if max_maturity < (maturity - data.access_time.date()).days / 365:
                        continue

            # Get calls and puts tables for this maturity and join them
            calls = yf_ticker.option_chain(maturity_str).calls
            calls['type'] = 'C'
            puts = yf_ticker.option_chain(maturity_str).puts
            puts['type'] = 'P'
            calls_and_puts = pd.concat([calls, puts], ignore_index=True)
            calls_and_puts['maturity'] = pd.Timestamp(maturity)
            calls_and_puts['price'] = (calls_and_puts['bid'] + calls_and_puts['ask']) / 2
            calls_and_puts['spread'] = calls_and_puts['ask'] - calls_and_puts['bid']
            # Add options for this maturity to the list, keeping only those with non-zero spread
            # (zero spread means no bid and ask prices)
            options_dataframes.append(calls_and_puts[calls_and_puts['spread'] > 0])

        # No options were loaded
        if not options_dataframes:
            raise RuntimeError('No options were loaded. '
                               'Check that the ticker and maturities are correct')

        # Make a single DataFrame which contains all the options. Keep only the columns we need.
        data.options = pd.concat(options_dataframes, ignore_index=True)[
                ['maturity', 'type', 'strike', 'price', 'spread', 'volume']]
        # Set multi-index and sort in the ascending order for maturities and strikes, and in the
        # descending order for type (i.e. 'P' before 'C' - this is convenient because when we
        # use only OTM options, puts will have smaller strikes, so we want them to be before calls)
        data.options.set_index(['maturity', 'type', 'strike'], inplace=True)
        data.options.sort_index(ascending=[True, False, True], inplace=True)

        # Make a DataFrame with forward prices, which will be filled later. Now it only contains
        # maturity and timeToMaturity columns
        # data.forward_prices = pd.DataFrame(
        #     data=[(maturity, time_to_maturities[maturity])
        #           for maturity in sorted(time_to_maturities)],
        #     columns=['maturity', 'timeToMaturity'])
        # data.forward_prices.set_index('maturity', inplace=True)

        return data
    
    def compute_forwards(self, discount_rate=0.0, monotone=False):
        """Computes forward prices from the options data.

        Args:
            discount_rates (float | callable): Risk-free rates to use for discounting. If a float,
                then a constant risk-free rate is assumed for all option maturities. If a callable,
                it must accept a single argument (time to maturity in years) and return the
                discount rate.
            monotone (bool): Ensure that the forward prices are non-decreasing with maturity by
                setting each price not smaller than the price for the previous maturity. This may
                break the call-put parity.

        Returns:
            The object itself with added forward prices and added columns `logStrike` and `OTM` to
            the options DataFrame.
        """
        if not hasattr(self, 'options'):
            raise RuntimeError('You need to load options data first')

        # Creating a dataframe and adding time to maturity (business and calendar)
        maturities = self.options.index.unique('maturity')
        self.forward_prices = pd.DataFrame(
             data=[(maturity, 
                    (maturity.date()-self.access_time.date()).days/365,
                    np.busday_count(self.access_time.date(), maturity.date())/250)
                   for maturity in maturities],
             columns=['maturity', 'calTimeToMaturity', 'busTimeToMaturity'])
        self.forward_prices.set_index('maturity', inplace=True)

        # Correct negative maturities (for options expiring on the same day as access_time)
        self.forward_prices.calTimeToMaturity[self.forward_prices.calTimeToMaturity<0] = 0
        self.forward_prices.busTimeToMaturity[self.forward_prices.busTimeToMaturity<0] = 0

        # Make a function for computing the appropriate discount factors
        if np.isscalar(discount_rate):
            def discount(t):
                return math.exp(-t*discount_rate)
        else:
            def discount(t):
                return math.exp(-t*discount_rate(t))

        # Fill in the discountRate column of forward_prices
        self.forward_prices['discountFactor'] = self.forward_prices.calTimeToMaturity.apply(discount)

        # A function for computing forward prices for each maturity, which will be called for each
        # row of the forward_prices DataFrame
        def compute_forward_price(forward):
            # maturity is the index of the forward_prices DataFrame, hence we use .name
            maturity = forward.name
            # Collect call and put options for the current maturity
            calls = self.options.loc[maturity].loc['C']
            puts = self.options.loc[maturity].loc['P']

            # By call-put parity, call and put prices must be equal for K=F. Hence as initial guess
            # for the forward price we take the strike where |C-P| is smallest
            # Note: we need to check for non-zero prices because some options may have zero prices
            # (e.g. if the volume is zero)
            price_diff = (calls.price[calls.price > 0] - puts.price[puts.price > 0]).dropna().abs()

            # Check that there are some options with non-zero prices and matching strikes
            # Typically, there will be no such options if we are trying to get data before the
            # market opens
            if not price_diff.empty:
                K = price_diff.idxmin()
                P = puts.loc[K].price
                C = calls.loc[K].price
                T = forward.calTimeToMaturity
                d = discount(T)
                return K + d*(C-P)
            else:
                return math.nan

        # Fill in the forwardPrice column of forward_prices
        self.forward_prices['forwardPrice'] = self.forward_prices.apply(compute_forward_price, axis=1)

        # Apply monotonicity correction
        if monotone:
            for i in range(1, len(self.forward_prices)):
                self.forward_prices.forwardPrice.iloc[i] = max(
                    self.forward_prices.forwardPrice.iloc[i], self.forward_prices.forwardPrice.iloc[i-1])
                           
        # Add logStrike and OTM columns to the options DataFrame
        self.options['logStrike'] = self.options.apply(
            # index: x.name = (maturity, type, strike)
            lambda x: math.log(x.name[2] / self.forward_prices.loc[x.name[0]].forwardPrice),
            axis=1)
        self.options['OTM'] = self.options.apply(
            lambda x: x.logStrike >= 0 if x.name[1] == 'C' else x.logStrike <= 0,
            axis=1)

        return self
    
    def compute_volatility(self, otm_vol_only=True):
        """Computes implied volatilities.

        Args:
            vol_type (str or list of str): Type of implied volatility to compute. Can be 'mid' for
                mid implied volatility, 'bid' for bid implied volatility, 'ask' for ask implied
                volatility, or a list of these.
            otm_vol_only (bool): If True, then implied volatilities are computed only for
                out-of-the-money and at-the-money options and set to NaN for in-the-money options.

        Returns:
            The object itself with added forward prices and implied volatilities.
        """
        # Define three functions for computing implied volatility, delta and total variance
        # which will be called for each row of the options DataFrame
        def compute_iv(option):     # option is a row of the options DataFrame
            if otm_vol_only and not option.OTM:
                return math.nan
            t = self.forward_prices.loc[option.name[0]].busTimeToMaturity
            if t == 0:
                return math.nan
            return implied_vol(
                forward_price=self.forward_prices.loc[option.name[0]].forwardPrice,
                maturity=t,
                strike=option.name[2],
                option_price=option.price,
                call_or_put_flag=(1 if option.name[1] == 'C' else -1),
                discount_factor=self.forward_prices.loc[option.name[0]].discountFactor)

        def compute_totalvar(option):
            if option.impliedVol is math.nan:
                return math.nan
            return option.impliedVol**2 * self.forward_prices.loc[option.name[0]].busTimeToMaturity
        
        def compute_delta(option):
            if option.impliedVol is math.nan:
                return math.nan
            d1 = 0.5*math.sqrt(option.totalVar) - option.logStrike / math.sqrt(option.totalVar)
            d = self.forward_prices.loc[option.name[0]].discountFactor
            T = self.forward_prices.loc[option.name[0]].busTimeToMaturity
            if option.name[1] == 'C':
                return d * norm.cdf(d1)
            else:
                return d * (norm.cdf(d1) - 1)

        self.options['impliedVol'] = self.options.apply(compute_iv, axis=1)
        self.options['totalVar'] = self.options.apply(compute_totalvar, axis=1)
        self.options['delta'] = self.options.apply(compute_delta, axis=1)

        return self

    def save_to_disk(self, filename, backend=None):
        """Saves data to disk.

        Supports pickle and Excel backends. The pickle backend saves the object as is. The Excel
        backend saves the object as an Excel file with three sheets: Info, Forwards and Options. The
        Info sheet contains `self.ticker`, `self.access_time` and `self.underlying_price` attributes
        of the object. The Forwards and Options sheets contain, respectively, `self.forward_prices`
        and 'self.options' DataFrames.

        Args:
            filename (str): Name of the file to save the object to.
            backend (str): Backend to use for saving the object. Must be 'pickle' or 'excel'. If
                None, the backend is determined from the filename extension: '.pkl' or '.dat' for
                pickle and  '.xlsx' for Excel.

        Returns:
            None
        """
        if not hasattr(self, 'options'):
            raise RuntimeError('Looks like the data has not been loaded yet - aborting')
        if backend is None:
            # Determine backend from filename extension: pickle or excel
            if filename.endswith('.pkl') or filename.endswith('.dat'):
                backend = 'pickle'
            elif filename.endswith('.xlsx'):
                backend = 'excel'
            else:
                raise ValueError('Unknown file extension')

        if backend == 'pickle':
            with open(filename, 'wb') as f:
                pickle.dump(self, f)
        elif backend == 'excel':
            # A DataFrame with one row for the Info sheet
            info = pd.DataFrame(
                data=[[self.ticker, self.access_time.strftime('%Y-%m-%d %H:%M:%S %z'),
                       self.underlying_price]],
                columns=['ticker', 'accessTime', 'underlyingPrice'])
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                info.to_excel(writer, sheet_name='Info', header=True, index=False)
                self.options.to_excel(writer, sheet_name='Options', header=True, index=True)
                if hasattr(self, 'forward_prices'):
                    self.forward_prices.to_excel(writer, sheet_name='Forwards', header=True, index=True)

            ##########################################
            # Nice visual formatiing of Excel sheets #
            ##########################################    
            W = load_workbook(filename)

            # Info sheet
            # Columns: A = ticker, B = accessTime, C = underlyingPrice
            W['Info'].column_dimensions['A'].width = len('ticker') + 2
            W['Info'].column_dimensions['B'].width = len('YYYY-MM-DD HH:MM:SS +HHMM') + 2
            W['Info'].column_dimensions['C'].width = len('underlyingPrice') + 2
            W['Info']['C2'].number_format = '0.0000'

            # Forwards sheet
            # Columns: A = maturity, B = timeToMaturity, C = discountRate, D = forwardPrice
            if hasattr(self, 'forward_prices'):
                W['Forwards'].column_dimensions['A'].width = len('YYYY-MM-DD') + 2
                for cell in W['Forwards']['A']:
                    cell.number_format = 'YYYY-MM-DD'
                for col in ['B', 'C', 'D']:
                    W['Forwards'].column_dimensions[col].width = 16
                    for cell in W['Forwards'][col]:
                        cell.number_format = '0.0000'
                W["Forwards"].freeze_panes = "A2"

            # Options sheet
            # Columns: A = maturity, B = type, C = strike, D = price, E = spread,
            # F = volume, G = logStrike, H = OTM, I = impliedVol, J = totalVar, K = delta
            W['Options'].column_dimensions['A'].width = len('YYYY-MM-DD') + 2
            W['Options'].column_dimensions['B'].width = len('type') + 2
            W['Options'].column_dimensions['C'].width = len('strike') + 2
            for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                W['Options'].column_dimensions[col].width = 16
            for c in W['Options']['A']:
                c.number_format = 'YYYY-MM-DD'
            for col in ['D', 'E', 'F', 'G', 'I', 'J', 'K']:
                for cell in W['Options'][col]:
                    cell.number_format = '0.0000'
            W["Options"].freeze_panes = "D2"

            W.save(filename)
        else:
            raise ValueError(f'Unknown backend: {backend}')

    @classmethod
    def load_from_disk(cls, filename, backend=None):
        """Loads data from disk.

        Supports pickle and Excel backends. See `save_to_disk` for details on the format of the
        saved data.

        Warning: this method does not check that the provided file contains correct (and safe) data.
        Always perform safety check for data from untrusted sources.

        Args:
            filename (str): Name of the file to load the object from.
            backend (str): Backend to use for loading the object. Must be 'pickle' or 'excel'. If
                None, the backend is determined from the filename extension: '.pkl' or '.dat' for
                pickle and  '.xlsx' for Excel.

        Returns:
            YFinanceData object
        """
        if backend is None:
            # Determine backend from filename extension: pickle or excel
            if filename.endswith('.pkl') or filename.endswith('.dat'):
                backend = 'pickle'
            elif filename.endswith('.xlsx'):
                backend = 'excel'
            else:
                raise ValueError('Unknown file extension')

        if backend == 'pickle':
            with open(filename, 'rb') as f:
                return pickle.load(f)
        elif backend == 'excel':
            data = cls()
            info = pd.read_excel(
                filename, sheet_name='Info', header=0, index_col=False)
            data.ticker = info.ticker[0]
            data.access_time = pd.Timestamp(
                datetime.datetime.strptime(info.accessTime[0], '%Y-%m-%d %H:%M:%S %z'))
            data.underlying_price = info.underlyingPrice[0]
            data.options = pd.read_excel(
                filename, sheet_name='Options', header=0, index_col=[0, 1, 2])
            # Try to load the Forwards sheet if it exists
            try:
                data.forward_prices = pd.read_excel(
                    filename, sheet_name='Forwards', header=0, index_col=0)
            except:
                pass
            return data
        else:
            raise ValueError(f'Unknown backend: {backend}')

    def get_maturities(self, min_maturity=None, max_maturity=None):
        """Returns the list of available maturities in the given range.

        Args:
            min_maturity (str | Timestamp): Return maturities not smaller than this value.
            max_maturity (str | Timestamp): Return maturities not greater than this value.

        Returns:
            A dictionary with available maturities as keys in the string format "YYY-MM-DD" and
            business time to maturity as values.
        """
        return {m.strftime("%Y-%m-%d") : self.forward_prices.loc[m].busTimeToMaturity
                for m in self.forward_prices[min_maturity:max_maturity].index}
    
    def get_time_to_maturity(self, maturity):
        """Returns the business time to maturity for a given maturity.

        Args:
            maturity (str | Timestamp): Maturity of the options.

        Returns:
            float: Business time to maturity in years.
        """
        return self.forward_prices.loc[maturity].busTimeToMaturity
    
    def get_smile(self, maturity, min_abs_delta=None, return_total_var=False, flatten=False,
                  unpack=False, max_points=None):
        """The implied volatility or total variance smile of OTM options for a given maturity.

        Args:
            maturity (str | datetime | Timestamp | list): Maturity of the options. If a string, the
                format must be accepted by Pandas. Several maturities can be provided as a list.
            min_abs_delta (float): Only options with the absolute value of delta not smaller than
                this number will be included in the curve. If None, all options are included.
            return_total_var (bool): If False, return the implied volatility curve in the
                coordinates (strike, implied_vol). If True, return the total variance curve in the
                coordinates (log_strike, total_var), where `log_strike = log(strike/forward)`.
            flatten (bool): If True, the result is returned as a tuple of four numpy arrays. If
                False, the result is returned as a list of tuples. See the Returns section for
                details.
            unpack (bool): Unpacks the result if `flatten` is False.
            max_points (int): Maximum number of strikes to include in the smile. Strikes are taken
                at regular intervals from the whole range of strikes. If None, all strikes are
                included. 

        Returns:
            If `flatten` is False and 'unpack` is False, a list of tuples, each containing two
                floats (forward price and time to maturity) and two numpy arrays (strikes and
                implied volatilities, or log-strikes and total variances).
            If `flatten` is False and 'unpack` is True, the result is unpacked as a tuple of four
                lists.
            If `flatten` is True, a tuple of four numpy arrays of the same length: forward prices,
                times to maturity, strikes or log-strikes, and implied volatilities or total
                variances. In this case `unpack` is ignored.
        """
        if min_abs_delta is None:
            min_abs_delta = 0

        if isinstance(maturity, list):
            maturities_list = maturity
        else:
            maturities_list = [maturity]

        # Collect the data for each maturity
        ret = []
        for m in maturities_list:
            # Forward price and time to maturity
            f = self.forward_prices.loc[m].forwardPrice
            t = self.forward_prices.loc[m].busTimeToMaturity
            # options with the given maturity and filtered for delta
            options = self.options.loc[m].copy().reset_index()
            filtered_options = options[options.OTM & (options.delta.abs() >= min_abs_delta)]
            # Take only max_points strikes
            if max_points is not None and len(filtered_options) > max_points:
                step = len(filtered_options) // max_points
                filtered_options = filtered_options.iloc[::step]

            if return_total_var:
                data = filtered_options.set_index('logStrike').sort_index().totalVar
            else:
                data = filtered_options.set_index('strike').sort_index().impliedVol
            
            ret.append((f, t, data.index.to_numpy(), data.to_numpy()))

        if flatten:
            f = np.concatenate([[f]*len(x) for f, _, x, _ in ret])
            t = np.concatenate([[t]*len(x) for _, t, x, _ in ret])
            x = np.concatenate([x for _, _, x, _ in ret])
            y = np.concatenate([y for _, _, _, y in ret])
            return f, t, x, y
        else:
            if isinstance(maturity, list):
                if unpack:
                    return tuple(zip(*ret))
                else:
                    return ret
            else:
                return ret[0]
        

class CMTData:
    """Class for loading the Treasury par yield curve (CMT rates) for a given date.

    Attributes:
        date (date): Date for which the data is loaded.
        maturities (array): Available maturities, counted in years after `date`. Currently, the
            maturities are hard-coded to 1 Mo, 3 Mo, 6 Mo, 1 Yr, 2 Yr, 3 Yr, 5 Yr, 7 Yr, 10 Yr,
            20 Yr, and 30 Yr.
        yield (array): Par yield rates for the available maturities, as a decimal.

    Methods:
        load_from_web: Loads data from the Treasury website.
        save_to_disk: Saves data to disk.
        load_from_disk: Loads data from disk.
        ytm: Computes the yield-to-maturity curve from the par yield curve.
    """

    date: datetime.date
    maturities: np.array
    par_yield: np.array

    @staticmethod
    def load_from_web(date=None):
        """Loads the Treasury par yield curve from the Treasury website.

        Args:
            date (str): Date for which the data is loaded in the YYYY-MM-DD format. If not provided,
            the current date is used.
        
        Returns:
            CMTData: The object itself with the loaded data.

        Notes:
            If the yield curve is not available for the requested date, the nearest available date
            is used.
        """
        data = CMTData()
        data.maturities = np.array([1/12, 3/12, 6/12, 1, 2, 3, 5, 7, 10, 20, 30])

        if date is None:
            date_ = datetime.datetime.now().date()
        else:
            date_ = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        
        year = date_.year
        url = f"https://home.treasury.gov/resource-center/data-chart-center/interest-rates/daily-treasury-rates.csv/{year}/all?type=daily_treasury_yield_curve&field_tdr_date_value={year}&page&_format=csv"
        df = pd.read_csv(url, parse_dates=['Date'], index_col='Date', date_format='%m/%d/%Y',
                         usecols=['Date', '1 Mo', '3 Mo', '6 Mo', '1 Yr', '2 Yr', '3 Yr', '5 Yr',
                                  '7 Yr', '10 Yr', '20 Yr', '30 Yr'])
        i = df.index.get_indexer([pd.Timestamp(date_)], method='nearest')[0]
        data.date = df.index[i].date()
        data.par_yield = df.iloc[i].to_numpy()/100

        return data

    def save_to_disk(self, filename):
        """Saves data to disk in Excel format.

        Args:
            filename (str): Name of the file.

        Returns:
            None
        """
        cmt = pd.DataFrame(
            data=[self.par_yield*100],
            index=[self.date],
            columns=['1 Mo', '3 Mo', '6 Mo', '1 Yr', '2 Yr', '3 Yr', '5 Yr', '7 Yr', '10 Yr',
                    '20 Yr', '30 Yr'])
        cmt.index.rename('Date', inplace=True)

        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            cmt.to_excel(writer, sheet_name='CMT', header=True, index=True)

        # Nice formatting
        W = load_workbook(filename)
        W['CMT'].column_dimensions['A'].width = len('YYYY-MM-DD') + 2
        for col in'BCDEFGHIJKL': 
            W['CMT'].column_dimensions[col].width = 7
        W.save(filename)

    @staticmethod
    def load_from_disk(filename):
        """Loads data from disk.

        Args:
            filename (str): Name of the file.

        Returns:
            CMTData object.
        """
        data = CMTData()
        df = pd.read_excel(filename, sheet_name='CMT', header=0, index_col='Date', 
                           usecols=['Date', '1 Mo', '3 Mo', '6 Mo', '1 Yr', '2 Yr', '3 Yr',
                                    '5 Yr', '7 Yr', '10 Yr', '20 Yr', '30 Yr'])
        data.date = df.index[0].date()
        data.maturities = np.array([1/12, 3/12, 6/12, 1, 2, 3, 5, 7, 10, 20, 30])
        data.par_yield = df.iloc[0].to_numpy()/100
        return data

    def ytm(self):
        """Computes the yield-to-maturity curve.

        Returns:
            An array of the same shape as `self.maturities` with yield-to-maturity for each
            available maturity.
        """
        ytm = np.empty_like(self.par_yield)

        # For 1, 3, 6 months we can ignore the coupons
        ytm[0] = 12*math.log(1+self.par_yield[0]/12)   # 1 month
        ytm[1] = 4*math.log(1+self.par_yield[1]/4)     # 3 months
        ytm[2] = 2*math.log(1+self.par_yield[2]/2)     # 6 months

        # Bootstrap the rest
        for i in range(3, len(self.maturities)):
            t = self.maturities[i]
            tprev = self.maturities[i-1]
            coupon = self.par_yield[i]/2
            # The next formula is due to the replication argument.
            # By definition of par yield, the price of a bond with par yield `r` and face value 1
            # is 1. 
            # Then we try to replicate the bond maturiting at `t` with the bond maturing at `tprev`
            # (they pay the same coupons until `tprev`) minus the discounted face value of the 
            # latter bond, and discounted coupons paid between `tprev` and `t`.
            price = lambda r: (
                1/self.par_yield[i-1]*2*coupon*(1 - math.exp(-ytm[i-1]*tprev)) 
                + math.exp(-r*t) 
                + coupon*sum(math.exp(-r*s) for s in np.arange(tprev+0.5, t+0.1, 0.5)) 
                - 1)
            res = root_scalar(f=price, method='Newton', x0=ytm[i-1])
            if res.converged:
                ytm[i] = res.root
            else:
                ytm[i] = math.nan

        return ytm
    
    def nelson_siegel_curve(self):
        return NelsonSiegel.calibrate(self.maturities, self.ytm())
    